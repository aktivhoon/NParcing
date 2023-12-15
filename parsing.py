from asyncore import write
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, DEFAULT_FONT
from datetime import datetime, timedelta

def what_day_is_it(date):
    days = ['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일']
    day = date.weekday()
    return days[day]

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def set_thick_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")
    upper_cell_num = cell_range.split(":")[0][1:]
    lower_cell_num = cell_range.split(":")[1][1:]
    ws['A'+upper_cell_num].border = Border(top=thick, left=thick, right=thin, bottom=thin)
    ws['A'+lower_cell_num].border = Border(top=thin, left=thick, right=thin, bottom=thick)
    ws['G'+upper_cell_num].border = Border(top=thick, left=thin, right=thick, bottom=thin)
    ws['G'+lower_cell_num].border = Border(top=thin, left=thin, right=thick, bottom=thick)
    for i in range(int(upper_cell_num)+1, int(lower_cell_num)):
        ws['A'+str(i)].border = Border(top=thin, left=thick, right=thin, bottom=thin)
        ws['G'+str(i)].border = Border(top=thin, left=thin, right=thick, bottom=thin)
    for j in 'BCDEF':
        ws[j+upper_cell_num].border = Border(top=thick, left=thin, right=thin, bottom=thin)
        ws[j+lower_cell_num].border = Border(top=thin, left=thin, right=thin, bottom=thick)

def convert_name(name):
    # check whether the last character is an alphabet
    if 'a' <= name[-1] <= 'z' or 'A' <= name[-1] <= 'Z':
        # check whether name itself is korean
        if not ('a' <= name[0] <= 'z' or 'A' <= name[0] <= 'Z'):
            while 'a' <= name[-1] <= 'z' or 'A' <= name[-1] <= 'Z':
                # remove the last character of the name (ex. 김영훈A -> 김영훈)
                name = name[:-1]
    if len(name) == 2:
        return name[0] + "O"
    elif len(name) > 2:
        return name[0] + "O"*(len(name)-2) + name[-1]

def make_space(ws, n, start_row, row_name) :
    if n == 0:
        return
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+n-1, end_column=1)
    ws.cell(start_row,1,row_name)
    ws.cell(start_row,1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    for i in range(start_row, start_row+n):
        ws.row_dimensions[i].height = 118.8
        for j in range(1, 13):
            if j == 6 or j ==7 : # 의뢰사유 및 평가내용, 진단치료결과는 왼쪽 정렬
                ws.cell(i,j).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
            elif j == 8 :  # 자살시도 N 채워두기 
                ws.cell(i,j).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                ws.cell(i,j,"N")
          #  elif j == 5 or j >=9 :
          #      ws.cell(i,j).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
          #      ws.cell(i,j,"'")
            else : # 나머지는 가운데 정렬
                ws.cell(i,j).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

def write_adm_dc(ws, adm, dc, *num, start_row=1, ward_name=None):
    if num != () :
        empty_room, n_man, n_woman = num
    altered_row = start_row
    if ward_name != "61병동" and ward_name != "62병동":
        start_row -= 1
        ws.merge_cells(start_row=altered_row, start_column=1, end_row=altered_row+max(1,len(adm))+max(1,len(dc)), end_column=1)
    else:
        ws.merge_cells(start_row=altered_row, start_column=1, end_row=altered_row+max(1,len(adm))+max(1,len(dc))+1, end_column=1)
    currentCell = ws.cell(altered_row, 1)
    currentCell.alignment = Alignment(horizontal='center', vertical='center')
    currentCell = ws.cell(altered_row, 2)
    currentCell.alignment = Alignment(horizontal='center', vertical='center')
    if ward_name == "61병동" or ward_name == "62병동":
        currentCell = ws.cell(altered_row, 3)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells(start_row=start_row+1, start_column=2, end_row=start_row+max(1,len(adm)), end_column=2)
    currentCell = ws.cell(start_row+1, 2)
    currentCell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=start_row+max(1,len(adm))+max(1,len(dc))+1, start_column=2, end_row=start_row+max(1,len(adm))+max(1,len(dc))+1, end_column=7)
    currentCell = ws.cell(start_row+max(1,len(adm))+max(1,len(dc))+1, 2)
    currentCell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=start_row+1+max(1,len(adm)), start_column=2, end_row=start_row+max(1,len(adm))+max(1,len(dc)), end_column=2)
    currentCell = ws.cell(start_row+1+max(1,len(adm)), 2)
    currentCell.alignment = Alignment(horizontal='center', vertical='center')

    if ward_name == "61병동" or ward_name == "62병동":
        ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=7)
    ws.cell(altered_row, 1, ward_name)
    
    if ward_name == "61병동" or ward_name == "62병동":
        ws.cell(start_row, 2, "공실수: "+empty_room)
        ws.cell(start_row, 3, "입원대기자수 : 남 "+n_man+" 여 "+n_woman)

    ws.cell(start_row+1, 2, "입원: " + str(len(adm)))
    ws.cell(start_row+max(1,len(adm))+max(1,len(dc))+1, 2, "특이사항 없음")
    for idx, element in enumerate(adm):
        ws.merge_cells(start_row=start_row+1+idx, start_column=3, end_row=start_row+1+idx, end_column=6)
        ws.cell(start_row+1+idx, 3, element['patient'])
        ws.cell(start_row+1+idx, 7, element['admission_type'])
    if len(adm) == 0:
        ws.merge_cells(start_row=start_row+1, start_column=3, end_row=start_row+1, end_column=6)

    ws.cell(start_row+1+max(1, len(adm)), 2, "퇴원: " + str(len(dc)))
    for idx, element in enumerate(dc):
        ws.merge_cells(start_row=start_row+1+max(1,len(adm))+idx, start_column=3, end_row=start_row+1+max(1,len(adm))+idx, end_column=6)
        ws.cell(start_row+1+max(1,len(adm))+idx, 3, element['patient'])
        ws.cell(start_row+1+max(1,len(adm))+idx, 7, element['discharge_type'])
    if len(dc) == 0:
        ws.merge_cells(start_row=start_row+1+max(1,len(adm)), start_column=3, end_row=start_row+1+max(1,len(adm)), end_column=6)

def generate_excel(input1, input2,_61_empty,_61_man,_61_woman,_62_empty,_62_man,_62_woman,yesterday_ped, yesterday_adult,today_day_adult,today_night_ped, today_night_adult):
    DEFAULT_FONT.sz = 9
    if (input1.getvalue() == '') : 
        df1 = pd.DataFrame()
    else :
        df1 = pd.read_csv(input1, sep="\t", header=0)
    adm_61 = []
    adm_62 = []
    adm_37 = []
    adm_121 = []
    if (input2.getvalue() == '') :
        df2 = pd.DataFrame()
    else :
        df2 = pd.read_csv(input2, sep="\t", header=0)
    dc_61 = []
    dc_62 = []
    dc_37 = []
    dc_121 = []
    for idx, row in df1.iterrows():
        data = {}
        data['patient'] = str(row[0]) + " " + convert_name(row[1]) + " " + row[2][0:-1] + " " + row[6] + " 교수님 " + row[12]
        data['admission_type'] = row[13].split("(")[0]
        if data['admission_type'] == "보호의무자에 의한 입원":
            data['admission_type'] = "보호입원"

        adm_ward = int(row[5].split("-")[0])

        if adm_ward == 61:
            adm_61.append(data)
        elif adm_ward == 62:
            adm_62.append(data)
        elif adm_ward == 37:
            adm_37.append(data)
        elif adm_ward == 121:
            adm_121.append(data)

    for idx, row in df2.iterrows():
        data = {}
        data['patient'] = str(row[0]) + " " + convert_name(row[1]) + " " + row[2][0:-1] + " " + row[6] + " 교수님 " + row[12]
        data['discharge_type'] = row[14]

        adm_ward = int(row[5].split("-")[0])
        if adm_ward == 61:
            dc_61.append(data)
        elif adm_ward == 62:
            dc_62.append(data)
        elif adm_ward == 37:
            dc_37.append(data)
        elif adm_ward == 121:
            dc_121.append(data)

    write_wb = Workbook()
    write_ws = write_wb.active
    
    start_yesterday_ped = 3
    start_yesterday_adult = start_yesterday_ped + int(yesterday_ped)
    start_today_day_adult = start_yesterday_adult + int(yesterday_adult)
    start_today_night_ped = start_today_day_adult + int(today_day_adult)
    start_today_night_adult = start_today_night_ped + int(today_night_ped)

    start_61 = start_today_night_adult+int(today_night_adult)+1
    start_62 = max(1,len(adm_61))+max(1,len(dc_61))+2+start_61
    start_37 = start_62+max(1,len(adm_62))+max(1,len(dc_62))+2
    start_121 = start_37+max(1,len(adm_37))+max(1,len(dc_37))+1
    start_opd = start_121+max(1, len(adm_121))+max(1,len(dc_121))+1

    # 주말 여부 확인
    yesterday = datetime.today() - timedelta(1)
    yesterday_is_weekend = False
    if what_day_is_it(yesterday) == "토요일" or what_day_is_it(yesterday) == "일요일" :
        yesterday_is_weekend = True 

    make_space(write_ws, int(yesterday_ped),start_row =start_yesterday_ped, row_name="전일\n소아당직")
    make_space(write_ws, int(yesterday_adult),start_row =start_yesterday_adult, row_name="전일\n성인당직")
    make_space(write_ws, int(today_day_adult),start_row =start_today_day_adult, row_name="성인\n낮당직")
    if yesterday_is_weekend == True :
        make_space(write_ws, int(today_night_ped),start_row =start_today_night_ped, row_name="소아\n응급실")
        make_space(write_ws, int(today_night_adult),start_row =start_today_night_adult, row_name="성인\n응급실")
    else :
        make_space(write_ws, int(today_night_ped),start_row =start_today_night_ped, row_name="소아\n밤당직")
        make_space(write_ws, int(today_night_adult),start_row =start_today_night_adult, row_name="성인\n밤당직")

    write_adm_dc(write_ws, adm_61, dc_61, _61_empty,_61_man,_61_woman, start_row=start_61, ward_name="61병동")
    write_adm_dc(write_ws, adm_62, dc_62, _62_empty,_62_man,_62_woman, start_row=start_62, ward_name="62병동")
    write_adm_dc(write_ws, adm_37, dc_37, start_row=start_37, ward_name="낮병원")
    write_adm_dc(write_ws, adm_121, dc_121, start_row=start_121, ward_name="특실")
    write_ws.cell(start_opd, 1, "외래")
    currentCell = write_ws.cell(start_opd, 1)
    currentCell.alignment = Alignment(horizontal='center', vertical='center')
    write_ws.merge_cells(start_row=start_opd, start_column=2, end_row=start_opd, end_column=7)
    write_ws.cell(start_opd, 2, "특이사항 없음")
    currentCell = write_ws.cell(start_opd, 2)
    currentCell.alignment = Alignment(horizontal='center', vertical='center')
    write_ws.row_dimensions[1].height = 16.5
    write_ws.row_dimensions[2].height = 39.6
    write_ws.column_dimensions['A'].width = 9.5
    write_ws.column_dimensions['B'].width = 9.38
    write_ws.column_dimensions['C'].width = 13
    write_ws.column_dimensions['D'].width = 9.38
    write_ws.column_dimensions['E'].width = 9
    write_ws.column_dimensions['F'].width = 55
    write_ws.column_dimensions['G'].width = 41
    write_ws.column_dimensions['H'].width = 9
    write_ws.column_dimensions['I'].width = 12.7
    write_ws.column_dimensions['J'].width = 12.2
    write_ws.column_dimensions['K'].width = 18
    write_ws.column_dimensions['L'].width = 12
    
    set_border(write_ws, 'A{}:L{}'.format(2,start_61-2))
    set_border(write_ws, 'A{}:G{}'.format(start_61,start_opd)) 
    set_thick_border(write_ws, 'A{}:G{}'.format(start_61, start_62-1))
    set_thick_border(write_ws, 'A{}:G{}'.format(start_62, start_37-1))
    set_thick_border(write_ws, 'A{}:G{}'.format(start_37, start_121-1))
    set_thick_border(write_ws, 'A{}:G{}'.format(start_121, start_opd-1))
    set_thick_border(write_ws, 'A{}:G{}'.format(start_opd, start_opd))
    
    write_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    yesterday = datetime.today() - timedelta(1)
    write_ws.cell(1, 1, str(yesterday.year)+"년 "+str(yesterday.month)+"월 "+str(yesterday.day)+"일 "+what_day_is_it(yesterday)+" 당직보고")
    currentCell = write_ws.cell(1, 1)
    currentCell.alignment = Alignment(vertical='center')
    write_ws.cell(2,2,"성명\n병록번호\n성별/나이")
    write_ws.cell(2,3,"지정의\n마지막 외래")
    write_ws.cell(2,4,"주소")
    write_ws.cell(2,5,"진료시간")
    write_ws.cell(2,6,"의뢰 사유 및 평가 내용")
    write_ws.cell(2,7,"진단\n치료결과")
    write_ws.cell(2,8,"자살시도\n여부")
    write_ws.cell(2,9,"ER 입실 시간")
    write_ws.cell(2,10,"NP 의뢰 시간")
    write_ws.cell(2,11,"NP 문제 해결 시간\n(회신 시간 기준")
    write_ws.cell(2,12,"ER 퇴실 시간")
    for i in range(2,13) :
        write_ws.cell(2,i).alignment = Alignment(horizontal = 'center', vertical='center', wrapText=True)
    filename = "당직보고_"+yesterday.strftime("%Y%m%d")+".xlsx"
    write_wb.save(filename)
    return filename